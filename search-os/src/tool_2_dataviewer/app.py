"""
Tool 2: Data Viewer - Aplicación principal
Visualizador de CSVs pesados con AgGrid virtualizado
"""
import streamlit as st
import sys
import re
import io
from pathlib import Path
from datetime import datetime

# Añadir src al path
sys.path.insert(0, str(Path(__file__).parent.parent))

from shared import data_manager
from tool_2_dataviewer.csv_loader import normalize_sabi_data
from tool_2_dataviewer.aggrid_config import render_grid
from tool_2_dataviewer.llm_factory import LLMFactory
from tool_2_dataviewer.ai_columns import enrich_column_batch
from tool_2_dataviewer.data_chat import render_data_chat
import pandas as pd
from openpyxl.utils import get_column_letter

# Configuración de página
st.set_page_config(
    page_title="Data Viewer - Search OS",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)


def clean_ai_score_columns(df: pd.DataFrame, schema: dict) -> tuple:
    """
    Limpia columnas IA que puedan tener HTML guardado, extrayendo solo el valor numérico.

    Args:
        df: DataFrame con datos
        schema: Schema del proyecto con definiciones de columnas

    Returns:
        Tupla (DataFrame limpio, bool indicando si hubo cambios)
    """
    if df.empty or 'custom_columns_definitions' not in schema:
        return df, False

    df_clean = df.copy()
    cleaned = False

    for col_name, col_config in schema['custom_columns_definitions'].items():
        if col_config.get('type') == 'ai_score' and col_name in df_clean.columns:
            def clean_score_value(val):
                """Extrae el valor numérico de un score, incluso si contiene HTML"""
                if pd.isna(val) or val == '' or val is None:
                    return pd.NA
                val_str = str(val)
                # Si contiene HTML, intentar extraer el número
                match = re.search(r'(\d+\.?\d*)/10|(\d+\.?\d*)', val_str)
                if match:
                    return float(match.group(1) or match.group(2))
                # Si no hay HTML, intentar parsear directamente
                try:
                    return float(val_str)
                except:
                    return pd.NA

            # Limpiar la columna
            original_values = df_clean[col_name].copy()
            df_clean[col_name] = df_clean[col_name].apply(clean_score_value)
            # Convertir a float64 para mantener consistencia
            df_clean[col_name] = pd.to_numeric(df_clean[col_name], errors='coerce')

            # Verificar si hubo cambios (comparar valores, no tipos)
            # Comparar después de convertir ambos a numérico
            original_numeric = pd.to_numeric(original_values, errors='coerce')
            if not original_numeric.equals(df_clean[col_name]):
                cleaned = True

    return df_clean, cleaned


def load_and_clean_master_data(project_id: str) -> pd.DataFrame:
    """
    Carga y limpia automáticamente los datos del proyecto.
    Limpia columnas IA que puedan tener HTML guardado.

    Args:
        project_id: ID del proyecto

    Returns:
        DataFrame limpio
    """
    df = data_manager.load_master_data(project_id)
    try:
        schema = data_manager.load_schema(project_id)
        df_clean, was_cleaned = clean_ai_score_columns(df, schema)
        if was_cleaned:
            # Guardar datos limpios
            data_manager.save_master_data(project_id, df_clean)
        return df_clean
    except:
        return df


def export_to_csv(df: pd.DataFrame, filename_prefix: str = "export") -> bytes:
    """
    Exporta DataFrame a CSV manteniendo formato y columnas IA.

    Args:
        df: DataFrame a exportar
        filename_prefix: Prefijo para el nombre del archivo

    Returns:
        Bytes del archivo CSV
    """
    # Crear buffer en memoria
    output = io.StringIO()
    # Exportar a CSV (sin índice, con encoding UTF-8)
    df.to_csv(output, index=False, encoding='utf-8')
    # Convertir a bytes
    return output.getvalue().encode('utf-8')


def export_to_excel(df: pd.DataFrame, filename_prefix: str = "export") -> bytes:
    """
    Exporta DataFrame a Excel con formato preservado.

    Args:
        df: DataFrame a exportar
        filename_prefix: Prefijo para el nombre del archivo

    Returns:
        Bytes del archivo Excel
    """
    # Crear buffer en memoria
    output = io.BytesIO()

    # Crear Excel writer con openpyxl
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Datos')

        # Obtener la hoja para aplicar formato
        worksheet = writer.sheets['Datos']

        # Ajustar ancho de columnas automáticamente
        for idx, col in enumerate(df.columns, start=1):
            # Calcular ancho basado en contenido
            max_length = max(
                df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                len(str(col))
            )
            # Limitar ancho máximo a 50 caracteres
            adjusted_width = min(max_length + 2, 50)
            # Usar get_column_letter para convertir índice a letra de columna (A, B, ..., Z, AA, AB, ...)
            column_letter = get_column_letter(idx)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    output.seek(0)
    return output.read()

# Inicializar session state
if 'project_id' not in st.session_state:
    st.session_state.project_id = None
if 'df_master' not in st.session_state:
    st.session_state.df_master = None
if 'selected_list_id' not in st.session_state:
    st.session_state.selected_list_id = 'inbox'
if 'show_add_column_modal' not in st.session_state:
    st.session_state.show_add_column_modal = False

if 'show_new_view_modal' not in st.session_state:
    st.session_state.show_new_view_modal = False

# ============================================================================
# SIDEBAR: Gestión de Proyectos
# ============================================================================
st.sidebar.title("📁 Proyectos")
st.sidebar.markdown("---")

# Listar proyectos existentes
projects = data_manager.list_projects()

if projects:
    st.sidebar.markdown("### Proyectos Existentes")

    # Crear lista de opciones con formato más amigable
    project_options = ["-- Seleccionar Proyecto --"] + projects

    selected_project_idx = 0
    if st.session_state.project_id and st.session_state.project_id in projects:
        selected_project_idx = projects.index(st.session_state.project_id) + 1

    selected_option = st.sidebar.selectbox(
        "Selecciona un proyecto:",
        options=project_options,
        index=selected_project_idx,
        key="project_selector"
    )

    if selected_option != "-- Seleccionar Proyecto --":
        if st.session_state.project_id != selected_option:
            st.session_state.project_id = selected_option
            # Cargar datos del proyecto
            try:
                st.session_state.df_master = load_and_clean_master_data(selected_option)
            except Exception as e:
                st.session_state.df_master = None
                st.sidebar.error(f"Error al cargar proyecto: {e}")
            st.rerun()
else:
    st.sidebar.info("No hay proyectos creados")

st.sidebar.markdown("---")

# Formulario para crear nuevo proyecto
st.sidebar.markdown("### Crear Nuevo Proyecto")

with st.sidebar.form("create_project_form"):
    new_project_name = st.text_input(
        "Nombre del Proyecto",
        placeholder="Ej: Sector Vitivinícola 2025",
        key="new_project_name"
    )
    create_submitted = st.form_submit_button("Crear Proyecto", use_container_width=True)

    if create_submitted and new_project_name:
        try:
            new_project_id = data_manager.create_project(new_project_name)
            st.sidebar.success(f"✅ Proyecto creado!")
            st.session_state.project_id = new_project_id
            st.session_state.df_master = None
            st.rerun()
        except Exception as e:
            st.sidebar.error(f"❌ Error: {e}")

# ============================================================================
# ÁREA PRINCIPAL
# ============================================================================

# Si no hay proyecto seleccionado, mostrar landing
if not st.session_state.project_id:
    st.title("📊 Data Viewer")
    st.markdown("### Visualizador de CSVs pesados con AgGrid virtualizado")

    st.markdown("---")

    col1, col2 = st.columns([2, 1])

    with col1:
        st.markdown("""
        #### 🚀 Bienvenido al Data Viewer

        Esta herramienta te permite:

        - **Gestionar proyectos** de análisis de datos
        - **Cargar archivos CSV/Excel** de SABI y normalizarlos automáticamente
        - **Visualizar datos** en tablas interactivas con AgGrid
        - **Organizar empresas** en listas (Inbox, Shortlist, Descartados)
        - **Añadir columnas personalizadas** y etiquetas

        #### 📝 Para empezar:

        1. Crea un nuevo proyecto desde el sidebar (izquierda) 👈
        2. O selecciona un proyecto existente
        3. Una vez dentro, podrás cargar tus datos CSV/Excel
        """)

    with col2:
        st.info("""
        **💡 Tip**

        Los proyectos se guardan en:
        `data/processed/{project_id}/`

        Cada proyecto tiene:
        - `master_data.parquet`
        - `schema_config.json`
        """)

    if projects:
        st.markdown("---")
        st.markdown("### 📁 Proyectos Disponibles")
        for project in projects:
            if st.button(f"📂 {project}", key=f"select_{project}", use_container_width=True):
                st.session_state.project_id = project
                try:
                    st.session_state.df_master = load_and_clean_master_data(project)
                except Exception as e:
                    st.session_state.df_master = None
                st.rerun()

# Si hay proyecto seleccionado, mostrar interfaz del proyecto
else:
    # Cargar schema del proyecto
    try:
        schema = data_manager.load_schema(st.session_state.project_id)
        project_name = schema.get('project_name', st.session_state.project_id)
    except:
        schema = None
        project_name = st.session_state.project_id

    st.title(f"📊 {project_name}")

    # ========================================================================
    # SECCIÓN: Carga de Archivos (TICKET-03)
    # ========================================================================
    if st.session_state.df_master is None or len(st.session_state.df_master) == 0:
        st.markdown("### 📤 Cargar Archivo CSV/Excel")
        st.info("💡 Sube un archivo de SABI (CSV o Excel) para empezar. El sistema normalizará automáticamente las columnas.")

        uploaded_file = st.file_uploader(
            "Selecciona un archivo",
            type=['csv', 'xlsx', 'xls'],
            key="file_uploader"
        )

        if uploaded_file is not None:
            with st.spinner("Procesando archivo..."):
                try:
                    # Normalizar datos SABI
                    df_normalized = normalize_sabi_data(uploaded_file)

                    # Guardar en el proyecto
                    data_manager.save_master_data(st.session_state.project_id, df_normalized)
                    st.session_state.df_master = df_normalized

                    st.success(f"✅ Archivo cargado correctamente: **{len(df_normalized)}** filas procesadas")
                    st.rerun()

                except Exception as e:
                    st.error(f"❌ Error al procesar archivo: {e}")
                    import traceback
                    with st.expander("Detalles del error"):
                        st.code(traceback.format_exc(), language="text")

    # ========================================================================
    # SECCIÓN: Gestión de Columnas Personalizadas (TICKET-07)
    # ========================================================================
    if st.session_state.df_master is not None and len(st.session_state.df_master) > 0:
        # Botón para añadir nueva columna
        col1, col2 = st.columns([1, 10])
        with col1:
            if st.button("➕ Columna", key="add_column_btn", use_container_width=True):
                st.session_state.show_add_column_modal = True

        # Modal para crear nueva columna
        if st.session_state.get('show_add_column_modal', False):
            with st.expander("➕ Añadir Nueva Columna", expanded=True):
                with st.form("add_column_form"):
                    col_name = st.text_input(
                        "Nombre de la Columna",
                        placeholder="Ej: Estado, Notas, Prioridad...",
                        help="Nombre único para la nueva columna"
                    )

                    col_type = st.radio(
                        "Tipo de Columna",
                        ["Texto Libre", "Etiqueta", "🤖 Columna IA"],
                        help="Texto Libre: permite escribir cualquier texto. Etiqueta: solo valores predefinidos. Columna IA: enriquece datos automáticamente con inteligencia artificial."
                    )

                    # Si es tipo "Etiqueta", mostrar input para opciones
                    options_input = None
                    if col_type == "Etiqueta":
                        options_input = st.text_input(
                            "Opciones (separadas por comas)",
                            placeholder="Ej: Pendiente, Contactado, Rechazado",
                            help="Escribe las opciones separadas por comas"
                        )

                    # Si es tipo "Columna IA", mostrar campos adicionales
                    ai_prompt = None
                    ai_model = None
                    ai_smart_context = True
                    if col_type == "🤖 Columna IA":
                        ai_prompt = st.text_area(
                            "Prompt / Instrucciones para la IA",
                            placeholder="Ej: Analiza la edad de los administradores y determina el riesgo de relevo generacional. Devuelve un score de 0-10 donde 10 es riesgo muy alto.",
                            help="Describe qué debe analizar la IA y cómo debe puntuar (0-10)",
                            height=100
                        )

                        # Selector de modelo
                        model_options = LLMFactory.list_available_models()
                        model_display = [f"{name}" for _, name in model_options]
                        model_keys = [key for key, _ in model_options]

                        model_selected_idx = 0  # Default: instant (Groq)
                        ai_model = st.selectbox(
                            "Modelo de IA",
                            options=model_display,
                            index=model_selected_idx,
                            help="⚡ Instantáneo: Groq (rápido) | 🏭 Batch: DeepInfra (económico) | 🧠 Complejo: GPT-4o (razonamiento) | 📚 Contexto Largo: Gemini (análisis extenso)"
                        )
                        # Obtener la key del modelo seleccionado
                        ai_model_key = model_keys[model_display.index(ai_model)]

                        ai_smart_context = st.checkbox(
                            "Auto-optimizar columnas (Smart Context)",
                            value=True,
                            help="Reduce consumo de tokens seleccionando solo las columnas relevantes (recomendado)"
                        )

                    col_submit1, col_submit2 = st.columns([1, 1])
                    with col_submit1:
                        submit = st.form_submit_button("✅ Crear Columna", use_container_width=True, type="primary")
                    with col_submit2:
                        cancel = st.form_submit_button("❌ Cancelar", use_container_width=True)

                    if cancel:
                        st.session_state.show_add_column_modal = False
                        st.rerun()

                    if submit:
                        if not col_name or not col_name.strip():
                            st.error("❌ El nombre de la columna es obligatorio")
                        else:
                            col_name_clean = col_name.strip()

                            # Verificar que no exista ya
                            if col_name_clean in st.session_state.df_master.columns:
                                st.error(f"❌ La columna '{col_name_clean}' ya existe")
                            else:
                                try:
                                    # Si es columna IA
                                    if col_type == "🤖 Columna IA":
                                        if not ai_prompt or not ai_prompt.strip():
                                            st.error("❌ El prompt de la columna IA es obligatorio")
                                        else:
                                            # Crear columna IA
                                            data_manager.add_ai_column(
                                                st.session_state.project_id,
                                                col_name_clean,
                                                ai_prompt.strip(),
                                                ai_model_key,
                                                ai_smart_context
                                            )

                                            # Recargar datos y schema
                                            st.session_state.df_master = load_and_clean_master_data(st.session_state.project_id)
                                            schema = data_manager.load_schema(st.session_state.project_id)

                                            st.success(f"✅ Columna IA '{col_name_clean}' creada correctamente")
                                            st.session_state.show_add_column_modal = False
                                            st.rerun()

                                    else:
                                        # Convertir tipo UI a tipo interno
                                        col_type_internal = 'text' if col_type == "Texto Libre" else 'single_select'

                                        # Procesar opciones si es etiqueta
                                        options_list = None
                                        if col_type_internal == 'single_select':
                                            if not options_input or not options_input.strip():
                                                st.error("❌ Las etiquetas requieren al menos una opción")
                                            else:
                                                # Separar por comas y limpiar
                                                options_list = [opt.strip() for opt in options_input.split(',') if opt.strip()]
                                                if len(options_list) == 0:
                                                    st.error("❌ Debe haber al menos una opción válida")
                                                else:
                                                    # Crear columna
                                                    data_manager.add_custom_column(
                                                        st.session_state.project_id,
                                                        col_name_clean,
                                                        col_type_internal,
                                                        options_list
                                                    )

                                                    # Recargar datos y schema
                                                    st.session_state.df_master = data_manager.load_master_data(st.session_state.project_id)
                                                    schema = data_manager.load_schema(st.session_state.project_id)

                                                    st.success(f"✅ Columna '{col_name_clean}' creada correctamente")
                                                    st.session_state.show_add_column_modal = False
                                                    st.rerun()
                                        else:
                                            # Crear columna tipo texto
                                            data_manager.add_custom_column(
                                                st.session_state.project_id,
                                                col_name_clean,
                                                col_type_internal,
                                                None
                                            )

                                            # Recargar datos y schema
                                            st.session_state.df_master = load_and_clean_master_data(st.session_state.project_id)
                                            schema = data_manager.load_schema(st.session_state.project_id)

                                            st.success(f"✅ Columna '{col_name_clean}' creada correctamente")
                                            st.session_state.show_add_column_modal = False
                                            st.rerun()

                                except ValueError as e:
                                    st.error(f"❌ Error: {e}")
                                except Exception as e:
                                    st.error(f"❌ Error inesperado: {e}")
                                    import traceback
                                    with st.expander("Detalles del error"):
                                        st.code(traceback.format_exc(), language="text")

    # ========================================================================
    # SECCIÓN: Navegación por Pestañas (TICKET-05) - CON VISTAS PERSONALIZADAS
    # ========================================================================
    if st.session_state.df_master is not None and len(st.session_state.df_master) > 0:
        # Cargar schema para obtener listas
        if schema:
            # Obtener todas las vistas (sistema + personalizadas)
            all_views = data_manager.get_all_views(st.session_state.project_id)
            
            # Botón para crear nueva vista
            col_new_view, _ = st.columns([1, 10])
            with col_new_view:
                if st.button("➕ Nueva Vista", key="btn_new_view", use_container_width=True):
                    st.session_state.show_new_view_modal = True
            
            # Modal para crear nueva vista
            if st.session_state.show_new_view_modal:
                with st.container():
                    st.markdown("### 📝 Crear Nueva Vista")
                    view_name = st.text_input("Nombre de la vista:", key="new_view_name")
                    view_icon = st.text_input("Icono (emoji):", value="📋", key="new_view_icon")
                    
                    # Seleccionar columnas visibles
                    available_cols = [col for col in st.session_state.df_master.columns 
                                    if col not in ['_uid', '_list_id']]
                    visible_cols = st.multiselect(
                        "Columnas visibles:",
                        options=available_cols,
                        default=available_cols[:10] if len(available_cols) > 10 else available_cols,
                        key="new_view_visible_cols"
                    )
                    
                    col_create, col_cancel = st.columns(2)
                    with col_create:
                        if st.button("✅ Crear Vista", key="btn_create_view", use_container_width=True, type="primary"):
                            if view_name and view_icon:
                                try:
                                    view_id = data_manager.create_custom_view(
                                        st.session_state.project_id,
                                        view_name,
                                        view_icon,
                                        visible_cols
                                    )
                                    st.success(f"✅ Vista '{view_name}' creada correctamente")
                                    st.session_state.show_new_view_modal = False
                                    # Actualizar lista de vistas
                                    all_views = data_manager.get_all_views(st.session_state.project_id)
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"❌ Error al crear vista: {e}")
                            else:
                                st.warning("⚠️ Por favor, completa el nombre y el icono")
                    
                    with col_cancel:
                        if st.button("❌ Cancelar", key="btn_cancel_view", use_container_width=True):
                            st.session_state.show_new_view_modal = False
                            st.rerun()
            
            # Crear tabs con todas las vistas
            tab_names = [view['name'] for view in all_views]
            tab_ids = [view['id'] for view in all_views]
            view_types = [view.get('type', 'system') for view in all_views]  # 'system' o 'custom'

            tabs = st.tabs(tab_names)

            # Renderizar contenido de cada tab
            for idx, (tab, view_id, view_type) in enumerate(zip(tabs, tab_ids, view_types)):
                with tab:
                    # Botón para eliminar vista (solo para vistas personalizadas)
                    if view_type == 'custom':
                        col_delete_view, _ = st.columns([1, 10])
                        with col_delete_view:
                            if st.button("🗑️ Eliminar Vista", key=f"btn_delete_view_{view_id}", use_container_width=True):
                                try:
                                    data_manager.delete_custom_view(st.session_state.project_id, view_id)
                                    st.success(f"✅ Vista eliminada correctamente")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"❌ Error al eliminar vista: {e}")
                    
                    # Obtener datos de la vista usando get_view_data
                    try:
                        df_view = data_manager.get_view_data(st.session_state.project_id, view_id)
                    except Exception as e:
                        st.error(f"❌ Error al cargar vista: {e}")
                        df_view = pd.DataFrame()

                    # ========================================================================
                    # FILTRAR FILAS VACÍAS (TICKET-09 Enhancement)
                    # ========================================================================
                    # Eliminar filas donde todas las columnas (excepto sistema) están vacías
                    # Esto asegura que el scroll solo llegue hasta la última fila con datos
                    # IMPORTANTE: JSON arrays (strings que empiezan con '[') se consideran datos válidos
                    if len(df_view) > 0:
                        # Columnas de sistema que NO debemos considerar al detectar filas vacías
                        system_cols = ['_uid', '_list_id']

                        # Columnas de datos (excluyendo sistema)
                        data_cols = [col for col in df_view.columns if col not in system_cols]

                        # Filtrar: mantener solo filas que tengan al menos una columna con datos válidos
                        if data_cols:
                            def has_valid_data(row):
                                """
                                Verifica si una fila tiene al menos un valor válido.
                                Considera válidos:
                                - Valores no nulos y no vacíos
                                - JSON arrays (strings que empiezan con '[')
                                """
                                for val in row:
                                    if pd.notna(val):
                                        val_str = str(val).strip()
                                        # Considerar válido si:
                                        # 1. No está vacío y no es una representación de NaN
                                        # 2. Es un JSON array (empieza con '[')
                                        if val_str not in ['', 'nan', 'None', 'NaN', 'NoneType']:
                                            # Si es un JSON array, también es válido
                                            if val_str.startswith('['):
                                                return True
                                            # Si tiene contenido, es válido
                                            if len(val_str) > 0:
                                                return True
                                return False

                            # Crear máscara: True si la fila tiene al menos un valor válido
                            mask = df_view[data_cols].apply(has_valid_data, axis=1)
                            df_view = df_view[mask].copy()

                    if len(df_view) > 0:
                        # ============================================================
                        # VERIFICAR FILTROS DEL CHAT ANTES DE RENDERIZAR GRID
                        # ============================================================
                        # Inicializar DataFrame filtrado del chat si no existe
                        chat_filtered_key = f'data_chat_filtered_df_{view_id}'
                        if chat_filtered_key not in st.session_state:
                            st.session_state[chat_filtered_key] = df_view.copy()

                        # Verificar si hay filtros activos del chat (comparar longitudes y contenido)
                        df_view_filtered = st.session_state[chat_filtered_key]
                        has_chat_filters = len(df_view_filtered) != len(df_view)

                        # Usar DataFrame filtrado si hay filtros activos
                        df_for_grid = df_view_filtered if has_chat_filters else df_view

                        # Barra superior con información y exportación
                        col_info, col_export_all = st.columns([3, 1])
                        with col_info:
                            if has_chat_filters:
                                st.info(f"📊 Mostrando **{len(df_for_grid)}** empresas (de {len(df_view)} totales)")
                            else:
                                st.info(f"📊 Mostrando **{len(df_for_grid)}** empresas en esta vista")

                        with col_export_all:
                            # Botón de exportación para toda la vista visible
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            filename_base = f"{st.session_state.project_id}_{view_id}_all_{timestamp}"

                            col_csv_all, col_xlsx_all = st.columns(2)

                            with col_csv_all:
                                csv_data_all = export_to_csv(df_for_grid, filename_base)
                                st.download_button(
                                    "📄 CSV",
                                    data=csv_data_all,
                                    file_name=f"{filename_base}.csv",
                                    mime="text/csv",
                                    key=f"download_csv_all_{view_id}",
                                    use_container_width=True
                                )

                            with col_xlsx_all:
                                try:
                                    xlsx_data_all = export_to_excel(df_for_grid, filename_base)
                                    st.download_button(
                                        "📊 Excel",
                                        data=xlsx_data_all,
                                        file_name=f"{filename_base}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key=f"download_xlsx_all_{view_id}",
                                        use_container_width=True
                                    )
                                except Exception as e:
                                    st.error(f"Error Excel: {e}")
                                    st.info("💡 Instala openpyxl: pip install openpyxl")

                        if has_chat_filters:
                            if st.button("🔄 Resetear Filtros del Chat", key=f"reset_chat_filters_{view_id}", use_container_width=True):
                                st.session_state[chat_filtered_key] = df_view.copy()
                                st.rerun()
                            st.warning(f"🔍 **Filtros del chat activos** - La tabla muestra solo las empresas que cumplen los criterios del chat")

                        # ============================================================
                        # SECCIÓN: AgGrid (TICKET-04) - CON DATOS FILTRADOS DEL CHAT
                        # ============================================================
                        custom_defs = schema.get('custom_columns_definitions', {})
                        grid_response = render_grid(
                            df_for_grid,  # Usar DataFrame filtrado del chat
                            custom_defs=custom_defs,
                            key=f"grid_{view_id}"
                        )

                        # ============================================================
                        # SECCIÓN: Persistir Cambios Editados en AgGrid (TICKET-07)
                        # ============================================================
                        # Si hay columnas personalizadas editables, mostrar botón de guardar
                        custom_cols = list(custom_defs.keys())
                        if custom_cols and grid_response and 'data' in grid_response:
                            edited_data = grid_response.get('data', [])

                            # Verificar si edited_data tiene contenido (puede ser lista o DataFrame)
                            has_data = False
                            if isinstance(edited_data, pd.DataFrame):
                                has_data = not edited_data.empty
                            elif isinstance(edited_data, list):
                                has_data = len(edited_data) > 0
                            else:
                                has_data = bool(edited_data)

                            if has_data:
                                # Botón para guardar cambios manualmente
                                if st.button("💾 Guardar Cambios en Columnas Personalizadas",
                                           key=f"save_custom_changes_{view_id}",
                                           use_container_width=True,
                                           type="primary"):
                                    try:
                                        # Convertir datos editados a DataFrame
                                        if isinstance(edited_data, pd.DataFrame):
                                            df_edited = edited_data
                                        else:
                                            df_edited = pd.DataFrame(edited_data)

                                        # Obtener UIDs de las filas editadas
                                        if '_uid' in df_edited.columns:
                                            edited_uids = df_edited['_uid'].tolist()

                                            # Actualizar solo las columnas personalizadas editadas
                                            for uid in edited_uids:
                                                row_edited = df_edited[df_edited['_uid'] == uid].iloc[0]
                                                row_idx = st.session_state.df_master[
                                                    st.session_state.df_master['_uid'] == uid
                                                ].index

                                                if len(row_idx) > 0:
                                                    for col in custom_cols:
                                                        if col in row_edited and col in st.session_state.df_master.columns:
                                                            new_value = row_edited[col]
                                                            # Convertir a string para compatibilidad con Parquet
                                                            if pd.notna(new_value):
                                                                st.session_state.df_master.loc[row_idx[0], col] = str(new_value)
                                                            else:
                                                                st.session_state.df_master.loc[row_idx[0], col] = ''

                                            # Guardar cambios en Parquet
                                            data_manager.save_master_data(
                                                st.session_state.project_id,
                                                st.session_state.df_master
                                            )

                                            st.success("✅ Cambios guardados correctamente")
                                            st.rerun()
                                    except Exception as e:
                                        st.error(f"❌ Error al guardar cambios: {e}")
                                        import traceback
                                        with st.expander("Detalles del error"):
                                            st.code(traceback.format_exc(), language="text")

                        # ============================================================
                        # SECCIÓN: Data Chat - Asistente de Datos (TICKET-15)
                        # El chat retorna el DataFrame filtrado que se usa en el siguiente rerun
                        # ============================================================
                        df_chat_filtered = render_data_chat(df_view, key_prefix=view_id)
                        # Actualizar session state con el DataFrame filtrado del chat
                        st.session_state[chat_filtered_key] = df_chat_filtered

                        # ============================================================
                        # SECCIÓN: Action Bar - Mover Filas (TICKET-06)
                        # ============================================================
                        # Manejar selección de filas de forma segura
                        # grid_response puede ser None o no tener 'selected_rows'
                        # selected_rows puede ser None, lista, o DataFrame
                        if grid_response:
                            selected_rows_raw = grid_response.get('selected_rows')

                            # Convertir a lista si es necesario
                            if selected_rows_raw is None:
                                selected_rows = []
                            elif isinstance(selected_rows_raw, pd.DataFrame):
                                # Si es DataFrame, convertir a lista de diccionarios
                                selected_rows = selected_rows_raw.to_dict('records')
                            elif isinstance(selected_rows_raw, list):
                                selected_rows = selected_rows_raw
                            else:
                                selected_rows = []

                            if len(selected_rows) > 0:
                                # ============================================================
                                # SECCIÓN: Action Bar Flotante Mejorado (TICKET-16)
                                # ============================================================
                                st.markdown("---")
                                # Barra destacada para acciones
                                st.markdown(f"### 🎯 {len(selected_rows)} fila(s) seleccionada(s)")

                                # Obtener UIDs de filas seleccionadas
                                selected_uids = []
                                for row in selected_rows:
                                    if isinstance(row, dict):
                                        if '_uid' in row:
                                            selected_uids.append(row['_uid'])
                                    elif hasattr(row, 'get'):
                                        uid_val = row.get('_uid') if '_uid' in row else None
                                        if uid_val is not None:
                                            selected_uids.append(uid_val)
                                    elif hasattr(row, '_uid'):
                                        selected_uids.append(row._uid)

                                # Organizar acciones en columnas
                                col_move, col_delete, col_export = st.columns([2, 1, 1])

                                with col_move:
                                    st.markdown("**📦 Mover a otra vista:**")
                                    # Obtener todas las vistas disponibles
                                    all_views_available = data_manager.get_all_views(st.session_state.project_id)
                                    
                                    # Crear opciones de destino (excluyendo la vista actual)
                                    target_view_options = {
                                        view['name']: (view['id'], view.get('type', 'system'))
                                        for view in all_views_available
                                        if view['id'] != view_id  # Excluir vista actual
                                    }

                                    if target_view_options:
                                        target_view_name = st.selectbox(
                                            "Selecciona vista destino:",
                                            options=list(target_view_options.keys()),
                                            key=f"move_target_{view_id}",
                                            label_visibility="collapsed"
                                        )

                                        if st.button(
                                            f"✅ Mover {len(selected_uids)} fila(s)",
                                            key=f"move_button_{view_id}",
                                            use_container_width=True,
                                            type="primary"
                                        ):
                                            try:
                                                target_view_id, target_view_type = target_view_options[target_view_name]
                                                
                                                # Si es una vista del sistema, actualizar _list_id
                                                if target_view_type == 'system':
                                                    st.session_state.df_master.loc[
                                                        st.session_state.df_master['_uid'].isin(selected_uids),
                                                        '_list_id'
                                                    ] = target_view_id
                                                    
                                                    # Guardar cambios
                                                    data_manager.save_master_data(
                                                        st.session_state.project_id,
                                                        st.session_state.df_master
                                                    )
                                                else:
                                                    # Si es una vista personalizada, añadir a rowIds
                                                    data_manager.add_rows_to_view(
                                                        st.session_state.project_id,
                                                        target_view_id,
                                                        selected_uids
                                                    )
                                                
                                                # Si la vista actual es personalizada, remover las filas
                                                if view_type == 'custom':
                                                    data_manager.remove_rows_from_view(
                                                        st.session_state.project_id,
                                                        view_id,
                                                        selected_uids
                                                    )
                                                
                                                st.success(f"✅ {len(selected_uids)} fila(s) movida(s) a {target_view_name}")
                                                st.rerun()

                                            except Exception as e:
                                                st.error(f"❌ Error al mover filas: {e}")
                                    else:
                                        st.info("No hay otras vistas disponibles")

                                with col_delete:
                                    st.markdown("**🗑️ Eliminar:**")
                                    if st.button(
                                        f"🗑️ Borrar {len(selected_uids)} fila(s)",
                                        key=f"delete_button_{view_id}",
                                        use_container_width=True,
                                        type="secondary"
                                    ):
                                        try:
                                            # Eliminar filas del master_data
                                            st.session_state.df_master = st.session_state.df_master[
                                                ~st.session_state.df_master['_uid'].isin(selected_uids)
                                            ]

                                            # Guardar cambios
                                            data_manager.save_master_data(
                                                st.session_state.project_id,
                                                st.session_state.df_master
                                            )

                                            st.success(f"✅ {len(selected_uids)} fila(s) eliminada(s)")
                                            st.rerun()

                                        except Exception as e:
                                            st.error(f"❌ Error al borrar filas: {e}")

                                with col_export:
                                    st.markdown("**💾 Exportar:**")

                                    # Preparar DataFrame para exportar (solo filas seleccionadas)
                                    df_to_export = st.session_state.df_master[
                                        st.session_state.df_master['_uid'].isin(selected_uids)
                                    ].copy()

                                    # Generar nombre de archivo con timestamp
                                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                                    filename_base = f"{st.session_state.project_id}_{view_id}_{timestamp}"

                                    col_csv, col_xlsx = st.columns(2)

                                    with col_csv:
                                        csv_data = export_to_csv(df_to_export, filename_base)
                                        st.download_button(
                                            "📄 CSV",
                                            data=csv_data,
                                            file_name=f"{filename_base}.csv",
                                            mime="text/csv",
                                            key=f"download_csv_{view_id}",
                                            use_container_width=True
                                        )

                                    with col_xlsx:
                                        try:
                                            xlsx_data = export_to_excel(df_to_export, filename_base)
                                            st.download_button(
                                                "📊 Excel",
                                                data=xlsx_data,
                                                file_name=f"{filename_base}.xlsx",
                                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                key=f"download_xlsx_{view_id}",
                                                use_container_width=True
                                            )
                                        except Exception as e:
                                            st.error(f"Error Excel: {e}")
                                            st.info("💡 Instala openpyxl: pip install openpyxl")

                                # ============================================================
                                # SECCIÓN: Enriquecimiento IA (TICKET-13)
                                # ============================================================
                                # Detectar columnas IA disponibles
                                ai_columns = [
                                    col_name for col_name, col_config in custom_defs.items()
                                    if col_config.get('type') == 'ai_score'
                                ]

                                if ai_columns:
                                    st.markdown("---")
                                    st.markdown("### 🤖 Enriquecimiento IA")

                                    selected_ai_col = st.selectbox(
                                        "Columna IA a enriquecer:",
                                        options=ai_columns,
                                        key=f"ai_column_selector_{view_id}",
                                        help="Selecciona la columna IA que quieres ejecutar sobre las filas seleccionadas"
                                    )

                                    col_ai1, col_ai2 = st.columns([1, 1])

                                    with col_ai1:
                                        if st.button(
                                            f"🚀 Ejecutar IA en {len(selected_uids)} fila(s) seleccionada(s)",
                                            key=f"enrich_selected_{view_id}",
                                            use_container_width=True,
                                            type="primary"
                                        ):
                                            try:
                                                # Obtener índices de las filas seleccionadas
                                                selected_indices = st.session_state.df_master[
                                                    st.session_state.df_master['_uid'].isin(selected_uids)
                                                ].index.tolist()

                                                # Crear elementos de progreso
                                                progress_bar = st.progress(0)
                                                status_text = st.empty()

                                                # Ejecutar enriquecimiento
                                                stats = enrich_column_batch(
                                                    st.session_state.project_id,
                                                    selected_ai_col,
                                                    row_indices=selected_indices,
                                                    progress_bar=progress_bar,
                                                    status_text=status_text
                                                )

                                                # Limpiar elementos de progreso
                                                progress_bar.empty()
                                                status_text.empty()

                                                # Mostrar resultados
                                                if stats['success'] > 0:
                                                    st.success(
                                                        f"✅ Enriquecimiento completado: "
                                                        f"{stats['success']} fila(s) procesada(s) correctamente"
                                                    )
                                                if stats['errors'] > 0:
                                                    st.warning(
                                                        f"⚠️ {stats['errors']} error(es) durante el procesamiento. "
                                                        f"Revisa la consola para más detalles."
                                                    )

                                                # Recargar datos
                                                st.session_state.df_master = load_and_clean_master_data(
                                                    st.session_state.project_id
                                                )
                                                st.rerun()

                                            except Exception as e:
                                                st.error(f"❌ Error al ejecutar enriquecimiento IA: {e}")
                                                import traceback
                                                with st.expander("Detalles del error"):
                                                    st.code(traceback.format_exc(), language="text")

                                    with col_ai2:
                                        if st.button(
                                            f"🚀 Ejecutar IA en TODAS las filas ({len(df_view)})",
                                            key=f"enrich_all_{view_id}",
                                            use_container_width=True,
                                            type="secondary"
                                        ):
                                            try:
                                                # Obtener índices de todas las filas visibles
                                                all_indices = df_view.index.tolist()

                                                # Crear elementos de progreso
                                                progress_bar = st.progress(0)
                                                status_text = st.empty()

                                                # Ejecutar enriquecimiento
                                                stats = enrich_column_batch(
                                                    st.session_state.project_id,
                                                    selected_ai_col,
                                                    row_indices=all_indices,
                                                    progress_bar=progress_bar,
                                                    status_text=status_text
                                                )

                                                # Limpiar elementos de progreso
                                                progress_bar.empty()
                                                status_text.empty()

                                                # Mostrar resultados
                                                if stats['success'] > 0:
                                                    st.success(
                                                        f"✅ Enriquecimiento completado: "
                                                        f"{stats['success']} fila(s) procesada(s) correctamente"
                                                    )
                                                if stats['errors'] > 0:
                                                    st.warning(
                                                        f"⚠️ {stats['errors']} error(es) durante el procesamiento. "
                                                        f"Revisa la consola para más detalles."
                                                    )

                                                # Recargar datos
                                                st.session_state.df_master = load_and_clean_master_data(
                                                    st.session_state.project_id
                                                )
                                                st.rerun()

                                            except Exception as e:
                                                st.error(f"❌ Error al ejecutar enriquecimiento IA: {e}")
                                                import traceback
                                                with st.expander("Detalles del error"):
                                                    st.code(traceback.format_exc(), language="text")

                                with col2:
                                    if st.button(
                                        "❌ Deseleccionar",
                                        key=f"deselect_{view_id}",
                                        use_container_width=True
                                    ):
                                        st.rerun()
                    else:
                        st.info(f"📭 Esta vista está vacía. Selecciona empresas de otras vistas y muévelas aquí.")

        # ========================================================================
        # SECCIÓN: Estadísticas del Proyecto
        # ========================================================================
        st.markdown("---")
        st.markdown("### 📊 Estadísticas del Proyecto")

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric("Total Empresas", len(st.session_state.df_master))

        with col2:
            if '_list_id' in st.session_state.df_master.columns:
                inbox_count = len(st.session_state.df_master[
                    st.session_state.df_master['_list_id'] == 'inbox'
                ])
                st.metric("En Inbox", inbox_count)

        with col3:
            if 'revenue' in st.session_state.df_master.columns:
                total_revenue = st.session_state.df_master['revenue'].sum()
                st.metric("Revenue Total", f"{total_revenue/1_000_000:.1f}M€" if total_revenue >= 1_000_000 else f"{total_revenue/1_000:.1f}K€")

        with col4:
            if 'ebitda' in st.session_state.df_master.columns:
                total_ebitda = st.session_state.df_master['ebitda'].sum()
                st.metric("EBITDA Total", f"{total_ebitda/1_000_000:.1f}M€" if total_ebitda >= 1_000_000 else f"{total_ebitda/1_000:.1f}K€")
